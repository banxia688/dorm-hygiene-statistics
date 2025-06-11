import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from info_extraction.college_info_extraction import get_college_name, college_dict
from info_extraction.hygiene_info_extraction import hygiene_data


def adjust_column_width(ws, lengths_tuple):
    """
    设置每一列的单元格宽度

    参数：worksheet, 每一列单元格长度元组
    """
    i = 0
    for col in ws.columns:
        column = col[0].column_letter  # 定位列
        ws.column_dimensions[column].width = lengths_tuple[i]  # 调整宽度
        i += 1


def add_borders(ws, style='thin'):
    """
    为 Excel 工作表中的所有有内容的单元格添加边框。

    参数:
    style (str): 边框样式，默认是 'thin'，可选 'thick', 'medium', 'dashed', 'dotted', 'double' 等。
    """
    # 定义边框样式
    border_style = Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style)
    )

    # 遍历所有单元格并添加边框
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border_style


# 定义数字序数映射表：汉字数字 -> 拉伯数字
chinese_to_number_map = {
    '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8,
    '九': 9, '十': 10, '十一': 11, '十二': 12, '十三': 13, '十四': 14,
    '十五': 15, '十六': 16, '十七': 17, '十八': 18
}
# 反向映射表：阿拉伯数字 -> 汉字数字
number_to_chinese_map = {v: k for k, v in chinese_to_number_map.items()}


def convert_chinese_to_number(nested_list):
    """
    将嵌套列表中所有子列表的第一个元素从 'x院' 转化为对应的阿拉伯数字
    """
    converted_list = []
    for entry in nested_list:
        college_name = entry[0].replace('院', '')  # 提取 '一' 到 '十八'
        college_number = chinese_to_number_map.get(college_name, 0)  # 转换为数字
        new_entry = [college_number] + entry[1:]  # 用数字替换原来的学院名称
        converted_list.append(new_entry)
    return converted_list


def convert_number_to_chinese(nested_list):
    """
    将嵌套列表中所有子列表的第一个元素从阿拉伯数字转换回 'x院'
    """
    converted_list = []
    for entry in nested_list:
        college_number = entry[0]
        college_name = number_to_chinese_map.get(college_number, '') + '院'  # 转换回汉字
        new_entry = [college_name] + entry[1:]  # 用汉字替换数字
        converted_list.append(new_entry)
    return converted_list


def custom_sort(data):
    def sort_key(entry) -> tuple:
        # entry 格式: [学院, 性别, [楼栋], [区域], [宿舍号], 备注]
        college_name = entry[0]  # 学院名称
        gender = 0 if entry[1] == '男' else 1  # 性别排序：先男后女

        # 提取楼栋号
        building = int(entry[2][0])

        # 提取区域号
        building_area = int(entry[2][1]) if len(entry[2]) > 1 else 0

        # 提取宿舍号
        room = int(entry[2][2]) if len(entry[2]) > 2 else 0

        # 返回排序键
        return college_name, gender, building, building_area, room

    return sorted(data, key=sort_key)


# Debug: 从 txt 文本提取出来的初步数据
# print('从txt文本提取出来的初步数据:')
# for entity in hygiene_data:
#     print(entity)
# print()

# 把学院数据转化为数字，以便排序
converted_data = convert_chinese_to_number(hygiene_data)
# 排序完后把数字重新转化为学院名称
sorted_data = convert_number_to_chinese(custom_sort(converted_data))

# print('排序后的列表：')
# for entity in sorted_data:
#     print(entity)
# print()

# 转换嵌套列表为平坦结构
flat_data = []
for entry in sorted_data:
    building = entry[2][0]  # 楼栋信息
    room_info = entry[2][1]  # 宿舍号信息
    if len(entry[2]) > 2:
        room_info = entry[2][1] + '-' + entry[2][2]

    remarks = entry[3]

    # 获取完整的学院名称
    full_college_name = get_college_name(college_dict, entry[0])

    # 构建平坦化的行数据
    row = [full_college_name, entry[1], int(building), room_info, remarks]
    flat_data.append(row)

# 创建 DataFrame
df = pd.DataFrame(flat_data, columns=["学院", "性别", "楼栋", "宿舍号", "备注"])

print('表格预览：')
print(df)
print()

# 添加“总计”列 并把 DataFrame 导出为 Excel 文件
df['总计'] = df.groupby('学院')['学院'].transform('count')
current_year = str(datetime.now().year)
month_abbr = datetime.now().strftime("%b")
output_path = ("files/dorm_hygiene_statistics/" + current_year + '/' + current_year + ' ' + month_abbr + " Dorm Hygiene Statistics.xlsx")
directory_path = "files/dorm_hygiene_statistics/" + str(current_year) + '/'

try:
    # If file exists, remove it
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"删除文件: {output_path}")

    # If the directory is empty, remove it
    if os.path.exists(directory_path) and not os.listdir(directory_path):  # Ensure the directory is empty
        os.rmdir(directory_path)
        print(f"删除空目录: {directory_path}")

    # Create the directory if it does not exist
    if not os.path.exists(directory_path):
        os.mkdir(directory_path)
        print(f"创建目录: {directory_path}")

    # Save the DataFrame to Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
except PermissionError:
    print(f"文件 {output_path} 正在被占用，请先关闭该文件后重试。")
    exit(0)
except Exception as e:
    print(f"发生错误: {e}")
    exit(0)

# 美化表格
wb = load_workbook(output_path)
ws = wb.active
# 加单元格边框
add_borders(ws)
# 合并相同的“学院”，“性别”，“楼栋“和“总计”单元格
# 循环初值
prior_academy = prior_gender = prior_building = None
# 跳过表头不处理，从第二行开始处理
no_header_row = 2
start_row_academy = start_row_gender = start_row_building = no_header_row

for row in range(no_header_row, ws.max_row + 1):
    current_academy = str(ws[f'A{row}'].value).strip()
    current_gender = str(ws[f'B{row}'].value).strip()
    current_building = str(ws[f'C{row}'].value).strip()

    print(
        f"Row {row}: prior_academy = {prior_academy}, current_Academy = {current_academy}, "
        f"prior_gender = {prior_gender}, current_gender = {current_gender}, "
        f"prior_building = {prior_building}, current_Building = {current_building}")

    # row 为遍历到的某一行，取值范围是 [2, ws.max_row-1]
    # 处理学院列
    if current_academy != prior_academy:
        if prior_academy is not None and row - 1 != start_row_academy:
            ws.merge_cells(start_row=start_row_academy, start_column=1, end_row=row - 1, end_column=1)
            ws.merge_cells(start_row=start_row_academy, start_column=6, end_row=row - 1, end_column=6)

        prior_academy = current_academy
        start_row_academy = row

        # 学院更改时，先检查一次前面所遍历到的性别行是否需要合并，以免出现学院更改时女生宿舍不能正确合并的情况
        if prior_gender is not None:
            # 当某个学院的性别多于一行时才进行合并
            if row - 1 != start_row_gender:
                ws.merge_cells(start_row=start_row_gender, start_column=2, end_row=row - 1, end_column=2)
                print(
                    f"Row {start_row_gender}-{row - 1}: academy = {prior_academy}, "
                    f"gender = {prior_gender}, ———————————————————————————————————————————————————————————————————————— specialGenderMerged!")

        # 学院更改时，先检查一次前面所遍历到的楼栋行是否需要合并，以免出现学院更改时楼栋不能正确合并的情况
        if prior_building is not None:
            # 当某个学院同性别的楼栋多于一行时才进行合并
            if row - 1 != start_row_building:
                ws.merge_cells(start_row=start_row_building, start_column=3, end_row=row - 1, end_column=3)

        # 学院发生更改时，重置性别与楼栋行
        prior_gender = prior_building = None
        start_row_gender = start_row_building = row

    # 处理性别列
    if current_gender != prior_gender:
        if prior_gender is not None and row - 1 != start_row_gender:
            ws.merge_cells(start_row=start_row_gender, start_column=2, end_row=row - 1, end_column=2)
            print(
                f"Row {start_row_gender}-{row - 1}: academy = {prior_academy}, "
                f"gender = {prior_gender}, ———————————————————————————————————————————————————————————————————————— normalGenderMerged!")

        prior_gender = current_gender
        start_row_gender = row

    # 处理楼栋列
    if current_building != prior_building:
        if prior_building is not None and row - 1 != start_row_building:
            ws.merge_cells(start_row=start_row_building, start_column=3, end_row=row - 1, end_column=3)
        prior_building = current_building
        start_row_building = row

# 合并最后一个学院的相关单元格
if prior_academy is not None:
    ws.merge_cells(start_row=start_row_academy, start_column=1, end_row=ws.max_row, end_column=1)
    ws.merge_cells(start_row=start_row_academy, start_column=6, end_row=ws.max_row, end_column=6)
if prior_gender is not None:
    ws.merge_cells(start_row=start_row_gender, start_column=2, end_row=ws.max_row, end_column=2)
if prior_building is not None:
    ws.merge_cells(start_row=start_row_building, start_column=3, end_row=ws.max_row, end_column=3)

# 设置所有单元格居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设定列宽数组
col_length_tuple = (22, 6.5, 6.0, 10.5, 16.0, 8.38)
# 调整列宽
adjust_column_width(ws, col_length_tuple)

# 保存更改
wb.save(output_path)
wb.close()

print(f"\n已成功生成{month_abbr}月份的宿舍卫生情况汇总表： {output_path}")
